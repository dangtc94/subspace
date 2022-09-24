from os.path import isfile
from openpyxl import load_workbook
import os
from pathlib import Path


dir_path = str(Path(__file__).parent.absolute())
print(dir_path)
wb = load_workbook(filename=dir_path+'\data.xlsx', data_only=True)

ws = wb['data']

startNo = ws['F2'].value
endNo = ws['G2'].value

for x in range(startNo, endNo):
  ws['F4'].value = x

  newFileName = dir_path + "\\noded" + str(x) + ".txt"
  f = open(newFileName, "w")

  data_rows = []
  for row in ws['G6':'G20']:
    data_cols = []
    for cell in row:
        f.write(cell.value + '\n')
  f.close()